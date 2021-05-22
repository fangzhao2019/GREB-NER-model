import re
import os
from openpyxl import load_workbook
import random

reg=r'\<[A-Z]{2,6}\>.+?\</[A-Z]{2,6}\>'
reg3=r'[0-9]{1,}\.?[0-9]{0,}'
ws=load_workbook('G:/论文实验——实体识别与链接/汽车命名实体识别/words.xlsx').active
wordSet=[]
for i in range(2,ws.max_row+1):
    #if ws.cell(row=i,column=4).value==1:
    wordSet.append(ws.cell(row=i,column=1).value)

def commentLoading(filename):
    dataSet=[]
    f=open('newDataSet2.ann',encoding='utf-8')
    for row in f.readlines():
        dataSet.append(row.strip())
    return dataSet

def myfind(x,y):
    return [ a for a in range(len(y)) if y[a:a+len(x)] == x]

def tagRecognition(comment):
    tagSet=[]
    for i in re.finditer(reg,comment):
        pat=comment[i.start():i.end()]
        if 'CAR' in pat:
            tagSet.append(['CAR',pat[5:len(pat)-6],i.start(),i.end()])
        if 'BRAND' in pat:
            tagSet.append(['BRAND',pat[7:len(pat)-8],i.start(),i.end()])
    return tagSet

def dataGeneration(comment,tagSet):
    data={}
    #先识别标签内容
    for tag in tagSet:
        t=tag[1]
        data[tag[2]+2+len(tag[0])]=[t[0],'B-%s'%tag[0]]
        if len(t)>1:
            for i in range(1,len(t)):
                data[tag[2]+2+len(tag[0])+i]=[t[i],'I-%s'%tag[0]]

    #再识别非标签内容
    startIndex=0
    for tag in tagSet:
        endIndex=tag[2]
        for i in range(startIndex,endIndex):
            data[i]=[comment[i],'O']
        startIndex=tag[3]
    for i in range(startIndex,len(comment)):
        data[i]=[comment[i],'O']
    return data

def dataSaving(data,filename):
    f=open(filename,'w',encoding='utf-8')
    for d in data:
        f.write('%s\t%s\n'%(d[1][0],d[1][1]))
    #f.write('\n')
    f.close()

def dataProcessing(data_num,commentSet,path):
    if not os.path.exists(path):
        os.mkdir(path)
    for comment in commentSet:
        comment=comment.upper()
        tagSet=tagRecognition(comment)
        data=dataGeneration(comment,tagSet)
        data=sorted(data.items(),key=lambda x:x[0])
        dataSaving(data,'%s/%d.txt'%(path,data_num))
        data_num+=1
    return data_num

def main():
    commentSet=commentLoading('newDataSet2.ann')
    random.shuffle(commentSet)

    trainComment=commentSet[:3000]
    devComment=commentSet[3000:3500]
    testComment=commentSet[3500:4000]
    
    data_num=1
    data_num=dataProcessing(data_num,trainComment,'G:/论文实验——实体识别与链接/汽车命名实体识别/data/trainData')
    data_num=dataProcessing(data_num,devComment,'G:/论文实验——实体识别与链接/汽车命名实体识别/data/devData')
    data_num=dataProcessing(data_num,testComment,'G:/论文实验——实体识别与链接/汽车命名实体识别/data/testData')
    
    
    

    
path='G:/论文实验——实体识别与链接/汽车命名实体识别/data'
comment='选车时看了<car>H6</car><car>宋max</car>和<car>rx5</car>，最后因为空间选择了<car>宋max</car>'
comment=comment.upper()
tagSet=tagRecognition(comment)
data=dataGeneration(comment,tagSet)
data=sorted(data.items(),key=lambda x:x[0])
dataSaving(data,'%s/%s.txt'%(path,'example'))
for d in data:print(d[1][0],'\t',d[1][1])

#main()  
