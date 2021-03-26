from __future__ import division
from openpyxl import load_workbook
import numpy as np

#这里构建的关系属于有向关系

#车与车之间的相关度
def brandDataLoading():
    wb=load_workbook('G:/论文实验——实体识别与链接/汽车命名实体识别/汽车知识库/newBrand.xlsx')
    ws=wb.active
    brandSet=[]
    brandWeight=[]
    for i in range(2,ws.max_row+1):
        brand=ws.cell(row=i,column=2).value
        weight=np.log(int(ws.cell(row=i,column=5).value))
        brandSet.append(brand)
        brandWeight.append(weight)
    return brandSet,brandWeight

def carDataLoading():
    wb=load_workbook('G:/论文实验——实体识别与链接/汽车命名实体识别/汽车知识库/newCar.xlsx')
    ws=wb.active
    brandCarSet={}
    carSet=[]
    configSet=[]
    weightSet=[]
    for i in range(2,ws.max_row+1):
        brand=ws.cell(row=i,column=1).value
        car=ws.cell(row=i,column=2).value
        length=ws.cell(row=i,column=4).value
        width=ws.cell(row=i,column=5).value
        height=ws.cell(row=i,column=6).value
        #价格数据取对数保证变化不至于过大
        avgPrice=ws.cell(row=i,column=8).value
        weight=np.log(ws.cell(row=i,column=12).value)
        if not brand in brandCarSet.keys():
            brandCarSet[brand]=[]
        brandCarSet[brand].append(car)
        carSet.append(car)
        configSet.append([length, width, height, avgPrice])
        weightSet.append(weight)
    return brandCarSet,carSet,configSet,weightSet

#车型-车型关系
def C2CrelationMeasure(car1,car2):
    #这里的体积差距取占主要车型的比例
    car1Size=np.sqrt(np.square(car1[0])+np.square(car1[1])+np.square(car1[2]))
    sizeRelation=np.sqrt(np.square(car1[0]-car2[0])+np.square(car1[1]-car2[1])+np.square(car1[2]-car2[2]))/car1Size
    if sizeRelation>1:sizeRelation=1
    #这里价格不取对数，而是取价格差占主要车型价格的比例
    priceRelation=(abs(car1[3]-car2[3])/car1[3])**2
    if priceRelation>1:priceRelation=1
    return sizeRelation,priceRelation

def C2CrelationMat(carSet,configSet):
    print('<车型-车型>关系矩阵生成中')
    sizeRelation=np.zeros((len(carSet),len(carSet)))
    priceRelation=np.zeros((len(carSet),len(carSet)))
    for i in range(len(carSet)):
        for j in range(len(carSet)):
            if i==j:continue
            sizeRelation[i,j],priceRelation[i,j]=C2CrelationMeasure(configSet[i],configSet[j])
    #sizeRelation=np.sqrt(sizeRelation)
    #这里按行取最大/小值后需要将其放大为原矩阵大小
    sizeRelationMax,sizeRelationMin=sizeRelation.max(axis=1),sizeRelation.min(axis=1)
    sizeRelationMax=np.array(np.matrix(np.tile(sizeRelationMax,(len(carSet),1))).transpose())
    sizeRelationMin=np.array(np.matrix(np.tile(sizeRelationMin,(len(carSet),1))).transpose())
    sizeRelation=(sizeRelationMax-sizeRelation)/(sizeRelationMax-sizeRelationMin)
    #priceRelation=np.sqrt(priceRelation)
    priceRelationMax,priceRelationMin=priceRelation.max(axis=1),priceRelation.min(axis=1)
    priceRelationMax=np.array(np.matrix(np.tile(priceRelationMax,(len(carSet),1))).transpose())
    priceRelationMin=np.array(np.matrix(np.tile(priceRelationMin,(len(carSet),1))).transpose())
    priceRelation=(priceRelationMax-priceRelation)/(priceRelationMax-priceRelationMin)
    return sizeRelation,priceRelation

#品牌-车型关系
def B2CrelationMat(brandSet,brandCarSet,carSet):
    print('<品牌-车型>关系矩阵生成中')
    B2Crelation=np.zeros((len(brandSet),len(carSet)))
    for i in range(len(brandSet)):
        brand=brandSet[i]
        if brand in brandCarSet.keys():
            for j in range(len(carSet)):
                car=carSet[j]
                if car in brandCarSet[brand]:
                    B2Crelation[i][j]=1
    return B2Crelation

#品牌与品牌之间
def B2BrelationMeasure(brand1Car,brand2Car,carSet,sizeRelation,priceRelation,weightSet):
    relationSet1=[]
    relationSet2=[]
    for car1 in brand1Car:
        for car2 in brand2Car:
            index1,index2=(carSet.index(car1),carSet.index(car2))
            #这里考虑车型间相互关系
            relation1=sizeRelation[index1,index2]
            relation2=priceRelation[index1,index2]
            weight=weightSet[index2]
            relationSet1.append(relation1*weight)
            relationSet2.append(relation2*weight)
    B2Brelation1=sum(relationSet1)/np.sqrt(len(relationSet1))
    B2Brelation2=sum(relationSet2)/np.sqrt(len(relationSet2))
    return B2Brelation1,B2Brelation2

def B2BrelationMat(brandSet,carSet,brandCarSet,sizeRelation,priceRelation,weightSet):
    print('<品牌-品牌>关系矩阵生成中')
    B2Brelation1=np.zeros((len(brandSet),len(brandSet)))
    B2Brelation2=np.zeros((len(brandSet),len(brandSet)))
    for i in range(len(brandSet)):
        for j in range(len(brandSet)):
            if i==j:
                continue
            brand1,brand2=(brandSet[i],brandSet[j])
            if brand1 in brandCarSet.keys() and brand2 in brandCarSet.keys():
                brand1Car,brand2Car=(brandCarSet[brand1],brandCarSet[brand2])
                B2Brelation1[i,j],B2Brelation2[i,j]=B2BrelationMeasure(brand1Car,brand2Car,carSet,sizeRelation,priceRelation,weightSet)
    #归一化
    B2Brelation1Max,B2Brelation1Min=B2Brelation1.max(axis=1),B2Brelation1.min(axis=1)
    B2Brelation1Max=np.array(np.matrix(np.tile(B2Brelation1Max,(len(brandSet),1))).transpose())
    B2Brelation1Min=np.array(np.matrix(np.tile(B2Brelation1Min,(len(brandSet),1))).transpose())
    B2Brelation1=(B2Brelation1-B2Brelation1Min)/(B2Brelation1Max-B2Brelation1Min+(B2Brelation1Max==0))
    #归一化
    B2Brelation2Max,B2Brelation2Min=B2Brelation2.max(axis=1),B2Brelation2.min(axis=1)
    B2Brelation2Max=np.array(np.matrix(np.tile(B2Brelation2Max,(len(brandSet),1))).transpose())
    B2Brelation2Min=np.array(np.matrix(np.tile(B2Brelation2Min,(len(brandSet),1))).transpose())
    B2Brelation2=(B2Brelation2-B2Brelation2Min)/(B2Brelation2Max-B2Brelation2Min+(B2Brelation2Max==0))
    return B2Brelation1,B2Brelation2

def getC2Crelation_cut(car1,car2,carSet,sizeRelation,priceRelation):
    index1=carSet.index(car1)
    index2=carSet.index(car2)
    return sizeRelation[index1,index2],priceRelation[index1,index2]
def getC2Crelation_final(car1,car2,carSet,relationMat):
    index1=carSet.index(car1)
    index2=carSet.index(car2)
    return relationMat[index1,index2]
def getC2C_most_similar(car,j,carSet,relationMat):
    index=carSet.index(car)
    value=relationMat[index]
    cars={}
    for i in range(len(carSet)):
        cars[carSet[i]]=value[i]
    for b in sorted(cars.items(),key=lambda a:a[1],reverse=True)[:j]:
        print(b)

def getB2Crelation(brand,car,brandSet,carSet,B2Crelation):
    indexB=brandSet.index(brand)
    indexC=carSet.index(car)
    return B2Crelation[indexB,indexC]

def getB2Brelation_cut(brand1,brand2,brandSet,B2Brelation1,B2Brelation2):
    index1=brandSet.index(brand1)
    index2=brandSet.index(brand2)
    return B2Brelation1[index1,index2],B2Brelation2[index1,index2]
def getB2Brelation_final(brand1,brand2,brandSet,relationMat):
    index1=brandSet.index(brand1)
    index2=brandSet.index(brand2)
    return relationMat[index1,index2]
def getB2B_most_similar(brand,j,brandSet,relationMat):
    index=brandSet.index(brand)
    value=relationMat[index]
    brands={}
    for i in range(len(brandSet)):
        brands[brandSet[i]]=value[i]
    for b in sorted(brands.items(),key=lambda a:a[1],reverse=True)[:j]:
        print(b)


def getBrandWeight(brand,brandSet,brandWeight):
    index=brandSet.index(brand)
    return brandWeight[index]

def getCarWeight(car,carSet,carWeight):
    index=carSet.index(car)
    return carWeight[index]
        
if __name__=='__main__':
    print('\n有向关系计算中......')  
    brandSet,brandWeight=brandDataLoading()
    brandCarSet,carSet,configSet,carWeight=carDataLoading()
    sizeRelation,priceRelation=C2CrelationMat(carSet,configSet)
    B2Crelation=B2CrelationMat(brandSet,brandCarSet,carSet)
    B2Brelation1,B2Brelation2=B2BrelationMat(brandSet,carSet,brandCarSet,sizeRelation,priceRelation,carWeight)


#测试
#getB2B_most_similar('奇瑞',20,brandSet,(B2Brelation1+B2Brelation2)/2.)
#getC2C_most_similar('哈弗H6',50,carSet,(sizeRelation+priceRelation)/2.)
