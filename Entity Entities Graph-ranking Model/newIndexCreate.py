from openpyxl import load_workbook
import time
import re
import unicodedata
from xpinyin import Pinyin
p = Pinyin()
reg=re.compile('[0-9]{1,}')
reg2=re.compile('[0-9A-Z\+]{1,}')
reg3=re.compile('[A-Z]{1,}')
ws1=load_workbook('G:/论文实验——实体识别与链接/汽车命名实体识别/words.xlsx').active
wordSet=[]
for i in range(2,ws1.max_row+1):
    wordSet.append(ws1.cell(row=i,column=1).value)

def _is_chinese_char(cp):
    """Checks whether CP is the codepoint of a CJK character."""
    if ((cp >= 0x4E00 and cp <= 0x9FFF) or  #
            (cp >= 0x3400 and cp <= 0x4DBF) or  #
            (cp >= 0x20000 and cp <= 0x2A6DF) or  #
            (cp >= 0x2A700 and cp <= 0x2B73F) or  #
            (cp >= 0x2B740 and cp <= 0x2B81F) or  #
            (cp >= 0x2B820 and cp <= 0x2CEAF) or
            (cp >= 0xF900 and cp <= 0xFAFF) or  #
            (cp >= 0x2F800 and cp <= 0x2FA1F)):  #
        return True

    return False

def word2pinyin(name):
    pinyin=[]
    for w in name:
        if _is_chinese_char(ord(w)):
            pinyin.append(p.get_pinyin(w, tone_marks='marks'))
        else:pinyin.append(w)
    return ''.join(pinyin)
    

def charAnalysis(brand, name, other_name):
    numReplace={'0':'零','1':'一','2':'二','3':'三','4':'四','5':'五','6':'六','7':'七','8':'八','9':'九'}
    brand=brand.upper().strip().replace(' ','')
    name=name.upper().strip().replace(' ','')
    other_name=other_name.split(',')
    name=unicodedata.normalize("NFD", name)
    indexs=[]
    #先加入全称及其拼音
    indexs.append(name)
    indexs.append(word2pinyin(name))
    for nn in other_name:
        nn=nn.upper().strip()
        indexs.append(nn)
        indexs.append(word2pinyin(nn))
    #不包含品牌时（添加品牌）；包含品牌时，去除品牌
    if not brand in name:
        newname=(brand+name)
        indexs.append(newname)
        indexs.append(word2pinyin(newname))
    else:
        newname=name.replace(brand,'')
        if len(newname)>1:
            indexs.append(newname)
            indexs.append(word2pinyin(newname))
    newBrand=brand.replace('汽车','')
    if not newBrand in name:
        newname=(newBrand+name)
        indexs.append(newname)
        indexs.append(word2pinyin(newname))
    else:
        newname=name.replace(newBrand,'')
        if len(newname)>1:
            indexs.append(newname)
            indexs.append(word2pinyin(newname))
    #字母数字组合
    for i in re.finditer(reg2,name):
        if i.start()==0 or i.end()==len(name):
            if (i.end()-i.start())>1:
                indexs.append(name[i.start():i.end()])
        if i.start()==0 and i.end()<len(name):#RAV4荣放
            indexs.append(name[i.end():])
        if i.end()==len(name)-1:#奔驰A级，奔驰GLA级
            indexs.append(name[i.start():])
            if (i.end()-i.start())==1:#宝马1系
                num=name[i.start():i.end()]
                if num in numReplace.keys():
                    indexs.append(name.replace(num,numReplace[num]))
                    indexs.append(name[i.start():].replace(num,numReplace[num]))
    #位于末尾的单数字
    for i in re.finditer(reg,name):
        if i.end()==len(name):
            if (i.end()-i.start())==1:
                indexs.append('小%s'%name[i.start()])
                indexs.append('小%s'%numReplace[name[i.start()]])
    indexs=list(set([row for row in indexs if len(row)>0]))
    for num in numReplace.keys():
        if num in indexs:
            indexs.remove(num)
    return indexs

def reverseIndex(indexSet):
    newIndexSet={}
    for k in indexSet.keys():
        for w in indexSet[k]:
            if not w in newIndexSet.keys():
                newIndexSet[w]=[]
            newIndexSet[w].append(k)
    return newIndexSet

def indexCreate_main(filename): 
    wb2=load_workbook(filename)
    ws2=wb2.active

    indexSet={}
    for i in range(2,ws2.max_row+1):
        brand=ws2.cell(row=i,column=1).value
        name=ws2.cell(row=i,column=2).value
        other_name=ws2.cell(row=i,column=3).value
        if brand==None:brand=''
        if other_name==None:other_name=''
        indexs=charAnalysis(brand,name,other_name)
        indexSet[name]=indexs
    newIndexSet=reverseIndex(indexSet)
    return newIndexSet

if __name__=='__main__':
    print('......正在构建索引库......')
    time1=time.time()
    carIndexSet=indexCreate_main('汽车知识库/newCar.xlsx')
    brandIndexSet=indexCreate_main('汽车知识库/newBrand.xlsx')
    time2=time.time()
    print('索引库构建完成，耗时%d秒'%(time2-time1))

#indexs,pinyins=charSplit('某瑞星途TXL','')
#print(indexs,pinyins)
